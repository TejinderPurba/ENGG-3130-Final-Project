# ENGG*3130 - Modeling Complex Systems
## Final Project

# JupyterHub Initialization
#%%capture cap_out --no-stderr\n",
#import sys
#!{sys.executable} -m pip install ortools
#import warnings
#warnings.simplefilter('ignore')

from __future__ import print_function
import io
import networkx as nx
import numpy as np
import json
import string
import pickle
import random
import matplotlib.pyplot as plt
from itertools import combinations 

# Import MapQuest API Libraries from https://github.com/MapQuest/directions-api-python-client somehow
from RouteOptions import RouteOptions
from RouteService import RouteService

# Import Excel libraries
from openpyxl import Workbook
from openpyxl import load_workbook

# Import Google VRP libraries
from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp

#%matplotlib inline

data_file = "1_02.xlsm" #Enter the data file for the day to schedule
driver = '1642'
sleeman_location = "551 Clair Rd W, Guelph, ON N1L 1E9"

options = RouteOptions()
service = RouteService('LcBh6g357kImNIm0k2aU5bkFG9MBxeHr')

def get_weighted_dist_matrix(locations, loads, factor_weight):

    route_matrix = service.routeMatrix(locations=locations, oneToMany=True)
    dist_matrix = route_matrix['distance']
    weighted_dist_matrix = dist_matrix

    i = 0;
    for load_weight in loads:
        weighted_dist_matrix[i] = dist_matrix[i]*(1/load_weight)*factor_weight
        i+=1
    return weighted_dist_matrix


def get_weighted_time_matrix(locations, loads, factor_weight):

    route_matrix = service.routeMatrix(locations=locations, oneToMany=True)

    time_matrix = route_matrix['time']
    weighted_time_matrix = time_matrix

    i = 0;
    for load_weight in loads:
        weighted_time_matrix[i] = time_matrix[i] * (1/load_weight) * factor_weight
        i+=1
    return weighted_time_matrix


def calc_edges(locations, loads=None, factor=None, factor_weight=None):
    """Calculate the edge weight"""

    if factor is None:
        routeMatrixRaw = service.routeMatrix(locations=locations, oneToMany=True)
        distMatrix = routeMatrixRaw['distance']
        timeMatrix = routeMatrixRaw['time']

        milesPerGallon = 120.0 # Should be around 12.0, but inflated by 10 for analysis purposes
        costPerGallon =  43.5 # Should be around 4.35 CAD per US gallon, but inflated by 10 for analysis purposes
        gasCostMatrixTemp = [i/milesPerGallon for i in distMatrix]
        gasCostMatrix = [i*costPerGallon for i in gasCostMatrixTemp]
        
        driverWage = 0.069444444 # Should be around 0.00694, but inflated by 10 for analysis purposes
        timeCostMatrix = [i*driverWage for i in timeMatrix]

        routeMatrix = [i + j for i, j in zip(gasCostMatrix, timeCostMatrix)]
  
    elif factor is 'dist':
        routeMatrix = get_weighted_dist_matrix(locations=locations, loads=loads, factor_weight=factor_weight)
    
    elif factor is 'time':
        routeMatrix = get_weighted_time_matrix(locations=locations, loads=loads, factor_weight=factor_weight)
    
    return routeMatrix
  
def create_graph_from_mapquest():

    wb = load_workbook(data_file, read_only=False, keep_vba=True)
    ws = wb.active
    deliveries = {}
    load_weights = {}

    # Parse the excel sheet data
    for row in range(1, len(ws['A'])+1):
        if ('delivery' in str(ws['J'+str(row)].value)):
            if str(ws['A'+str(row)].value) in deliveries.keys():
                deliveries[str(ws['A'+str(row)].value)].append(str(ws['R'+str(row)].value).replace('#',''))
                if (ws['AK'+str(row)].value is not None):
                    load_weights[float(ws['A'+str(row)].value)].append(float(ws['AK'+str(row)].value))
                else:
                    load_weights[float(ws['A'+str(row)].value)].append(float(0.0))
            else:
                deliveries[str(ws['A'+str(row)].value)] = [str(ws['R'+str(row)].value).replace('#','')]
                if (ws['AK'+str(row)].value is not None):
                    load_weights[float(ws['A'+str(row)].value)] = [float(ws['AK'+str(row)].value)]
                else:
                    load_weights[float(ws['A'+str(row)].value)] = [float(0.0)]

    G = nx.DiGraph()
    drive_times_all = {}
    drive_times = {}
    
    ####### Single Driver Parsing #######
    #nodes = deliveries[driver] 

    ####### All Driver Parsing #######
    nodes = [y for x in deliveries.values() for y in x]

    nodes.insert(0, sleeman_location) # Add the origin destination

    for n in range(0, len(nodes)): 
        cycled_nodes = (nodes[-n:] + nodes[:-n])
        routeMatrix = calc_edges(cycled_nodes)

        for i in range(1, len(cycled_nodes)):
            #print cycled_nodes[0]+' ~ AND ~ '+cycled_nodes[i]
            #import ipdb
            #ipdb.set_trace()
            #print str(cycled_nodes[0]) +" --------- "+ str(cycled_nodes[i]) +" --------- "+ str(routeMatrix[i])
            drive_times_all[(cycled_nodes[0], cycled_nodes[i])] = float(routeMatrix[i])

    all_combinations = list(combinations(nodes, 2))
    for n in range(len(all_combinations)):
        val1 = drive_times_all[all_combinations[n]]
        reversed_combination = all_combinations[n][::-1]
        val2 = drive_times_all[reversed_combination]
        avg = (val1+val2)/2.0
        drive_times[all_combinations[n]] = round(avg, 1)
        drive_times[reversed_combination] = round(avg, 1)

    G.add_nodes_from(nodes)
    G.add_edges_from(drive_times)

    return G, drive_times, nodes

def display_graph(G, edges):
    nx.draw(G,
        node_color='b',
        node_size=2000,
        with_labels=True,
        font_size=10)
    nx.draw_networkx_edge_labels(G, edge_labels=edges, pos=nx.spring_layout(G), font_size=8)
    plt.show()

def create_data_model(num_vehicles=1):
    """Stores the data for the problem."""
    data = {}
    distance_matrix = []
    distance_matrix_list = []

    graphs_and_edges = create_graph_from_mapquest()
    drive_times = graphs_and_edges[1]
    delivery_nodes = graphs_and_edges[2]

    for n in range(len(delivery_nodes)-1):
        del distance_matrix_list[:]
        for i in range(len(delivery_nodes)-1):
            search_tuple = (delivery_nodes[n], delivery_nodes[i])
            if search_tuple in drive_times:
                distance_matrix_list.append(int(drive_times[search_tuple]))
            else:
                distance_matrix_list.append(int(0))
        #print distance_matrix_list # Debugging line
        distance_matrix.append(list(distance_matrix_list))

    data['distance_matrix'] = distance_matrix
    data['num_vehicles'] = num_vehicles 
    data['depot'] = 0 # Should be static?

    # Save dictionary to pickle file so dont have to use MapQuest API every time
    file_name = 'CostMatrix-%s-%s.pickle' % (str(data_file[-9:-5]), str(num_vehicles))
    with open(file_name, 'wb') as handle:
        pickle.dump(data, handle, protocol=pickle.HIGHEST_PROTOCOL)

    return data

"""https://developers.google.com/optimization/routing/tsp"""
def print_solution_tsp(manager, routing, solution):
    """Prints solution on console."""
    index = routing.Start(0)
    plan_output = 'Route for vehicle 0:\n'
    route_distance = 0
    while not routing.IsEnd(index):
        plan_output += ' {} ->'.format(manager.IndexToNode(index))
        previous_index = index
        index = solution.Value(routing.NextVar(index))
        route_distance += routing.GetArcCostForVehicle(previous_index, index, 0)
    plan_output += ' {}\n'.format(manager.IndexToNode(index))
    plan_output += '\nRoute cost: {} units\n'.format(route_distance)
    print(plan_output)

"""https://developers.google.com/optimization/routing/vrp"""
def print_solution_vrp(data, manager, routing, solution):
    """Prints solution on console."""
    max_route_distance = 0
    total_distance = 0
    for vehicle_id in range(data['num_vehicles']):
        index = routing.Start(vehicle_id)
        plan_output = 'Route for vehicle {}:\n'.format(vehicle_id)
        route_distance = 0
        while not routing.IsEnd(index):
            plan_output += ' {} -> '.format(manager.IndexToNode(index))
            previous_index = index
            index = solution.Value(routing.NextVar(index))
            route_distance += routing.GetArcCostForVehicle(
                previous_index, index, vehicle_id)
        plan_output += '{}\n'.format(manager.IndexToNode(index))
        plan_output += 'Cost of the route: {} units\n'.format(route_distance)
        print(plan_output)
        total_distance += route_distance
        max_route_distance = max(route_distance, max_route_distance)
    print('Total cost of all routes: {} units'.format(total_distance))

def print_solution_vrp_limited(data, manager, routing, solution):
    """Prints solution on console."""

    total_distance = 0
    total_load = 0
    for vehicle_id in range(data['num_vehicles']):
        index = routing.Start(vehicle_id)
        plan_output = 'Route for vehicle {}:\n'.format(vehicle_id)
        route_distance = 0
        route_load = 0
        while not routing.IsEnd(index):
            node_index = manager.IndexToNode(index)
            route_load += data['demands'][node_index]
            plan_output += ' {0} Load({1}) -> '.format(node_index, route_load)
            previous_index = index
            index = solution.Value(routing.NextVar(index))
            route_distance += routing.GetArcCostForVehicle(
                previous_index, index, vehicle_id)
        plan_output += ' {0} Load({1})\n'.format(manager.IndexToNode(index),
                                                 route_load)
        plan_output += 'Cost of the route: {} units\n'.format(route_distance)
        plan_output += 'Load of the route: {}\n'.format(route_load)
        print(plan_output)
        total_distance += route_distance
        total_load += route_load
    print('Total cost of all routes: {} units'.format(total_distance))
    print('Total load of all routes: {}'.format(total_load))

"""https://developers.google.com/optimization/routing/tsp"""
def tsp_processing(data=None):

    # Create the routing index manager.
    manager = pywrapcp.RoutingIndexManager(len(data['distance_matrix']),
                                           data['num_vehicles'], data['depot'])

    # Create Routing Model.
    routing = pywrapcp.RoutingModel(manager)


    def distance_callback(from_index, to_index):
        """Returns the distance between the two nodes."""
        # Convert from routing variable Index to distance matrix NodeIndex.
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return data['distance_matrix'][from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)

    # Define cost of each arc.
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    # Setting first solution heuristic.
    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.local_search_metaheuristic = (routing_enums_pb2.LocalSearchMetaheuristic.GREEDY_DESCENT)
    search_parameters.time_limit.seconds = 30
    #search_parameters.log_search = True
    search_parameters.first_solution_strategy = (routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC) # FIRST SOLUTION STRATEGY WILL BE USED TO COMPARE ALGORITHMS

    # Solve the problem.
    solution = routing.SolveWithParameters(search_parameters)

    # Print solution on console.
    if solution:
        print_solution_tsp(manager, routing, solution)

"""https://developers.google.com/optimization/routing/vrp"""
def vrp_processing(data=None):                                                 

    # Create the routing index manager.
    manager = pywrapcp.RoutingIndexManager(len(data['distance_matrix']),
                                           data['num_vehicles'], data['depot'])

    # Create Routing Model.
    routing = pywrapcp.RoutingModel(manager)


    # Create and register a transit callback.
    def distance_callback(from_index, to_index):
        """Returns the distance between the two nodes."""
        # Convert from routing variable Index to distance matrix NodeIndex.
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return data['distance_matrix'][from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)

    # Define cost of each arc.
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

    # Add Distance constraint.
    dimension_name = 'Distance'
    routing.AddDimension(
        transit_callback_index,
        0,  # no slack
        5000,  # vehicle maximum travel distance
        True,  # start cumul to zero
        dimension_name)
    distance_dimension = routing.GetDimensionOrDie(dimension_name)
    distance_dimension.SetGlobalSpanCostCoefficient(100)

    # Setting first solution heuristic.

    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.local_search_metaheuristic = (routing_enums_pb2.LocalSearchMetaheuristic.GREEDY_DESCENT)
    search_parameters.time_limit.seconds = 30
    #search_parameters.log_search = True
    search_parameters.first_solution_strategy = (routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC) # FIRST SOLUTION STRATEGY WILL BE USED TO COMPARE ALGORITHMS

    # Solve the problem.
    solution = routing.SolveWithParameters(search_parameters)

    # Print solution on console.
    if solution:
        print_solution_vrp(data, manager, routing, solution)

"""https://developers.google.com/optimization/routing/cvrp"""
def vrp_limited_processing(data=None):

    manager = pywrapcp.RoutingIndexManager(len(data['distance_matrix']), data['num_vehicles'], data['depot'])

    # Create Routing Model.
    routing = pywrapcp.RoutingModel(manager)


    # Create and register a transit callback.
    def distance_callback(from_index, to_index):
        """Returns the distance between the two nodes."""
        # Convert from routing variable Index to distance matrix NodeIndex.
        from_node = manager.IndexToNode(from_index)
        to_node = manager.IndexToNode(to_index)
        return data['distance_matrix'][from_node][to_node]

    transit_callback_index = routing.RegisterTransitCallback(distance_callback)

    # Define cost of each arc.
    routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)


    # Add Capacity constraint.
    def demand_callback(from_index):
        """Returns the demand of the node."""
        # Convert from routing variable Index to demands NodeIndex.
        from_node = manager.IndexToNode(from_index)
        return data['demands'][from_node]

    demand_callback_index = routing.RegisterUnaryTransitCallback(demand_callback)
    routing.AddDimensionWithVehicleCapacity(
        demand_callback_index,
        0,  # null capacity slack
        data['vehicle_capacities'],  # vehicle maximum capacities
        True,  # start cumul to zero
        'Capacity')

    # Setting first solution heuristic.
    search_parameters = pywrapcp.DefaultRoutingSearchParameters()
    search_parameters.local_search_metaheuristic = (routing_enums_pb2.LocalSearchMetaheuristic.GREEDY_DESCENT)
    search_parameters.time_limit.seconds = 30
    #search_parameters.log_search = True
    search_parameters.first_solution_strategy = (routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC) # FIRST SOLUTION STRATEGY WILL BE USED TO COMPARE ALGORITHMS

    # Solve the problem.
    solution = routing.SolveWithParameters(search_parameters)

    # Print solution on console.
    if solution:
        print_solution_vrp_limited(data, manager, routing, solution)
        graph_limited_vrp(data, manager, routing, solution)

def loadPickle(fileName=None):
    with open(fileName, 'rb') as handle:
        data = pickle.load(handle)

    return data


def graph_limited_vrp(data, manager, routing, solution):
    """This function creates and displays a graph for the limited VRP model.
    This could also be used for the other models."""
    
    current_plot = 0
    plot_num = 0
    fig = plt.figure(figsize=(20, 8)) # create a figure to display
    empty_routes = 0

    # Get number of vehicles with no routes
    for vehicle_id in range(data['num_vehicles']):
        index = routing.Start(vehicle_id)
        node_index = 0
        while not routing.IsEnd(index):

            previous_index = node_index
            index = solution.Value(routing.NextVar(index))
            node_index = manager.IndexToNode(index) # get index of next node in path

            if previous_index is node_index:  # skips routes with ZERO deliveries
                empty_routes += 1
                break
    
    plot_num = data['num_vehicles'] - empty_routes # number of columns to plot

    # iterate through vehicles. Unique graph is created for every vehicle with >0 drops
    for vehicle_id in range(data['num_vehicles']):
        index = routing.Start(vehicle_id)
        G = nx.DiGraph() # initialize a graph for vehicle (vehicle_id)
        edges = {}
        node_index = 0

        while not routing.IsEnd(index):

            previous_index = node_index
            index = solution.Value(routing.NextVar(index))
            node_index = manager.IndexToNode(index) # get index of next node in path

            if previous_index is node_index:  # skips routes with ZERO deliveries
                break

            if node_index >= len(data['distance_matrix']):
                node_index = 0

            # This dictionary is later used to label the edges
            edges[(previous_index, node_index)] = routing.GetArcCostForVehicle(previous_index, index, vehicle_id)

            # Edges are added one by one, with length proportional to cost IF spring layout is chosen
            G.add_edge(previous_index, node_index, #length=routing.GetArcCostForVehicle(previous_index, index, vehicle_id),
                       color=colour_finder(vehicle_id))
            G.add_node(node_index)

        # IF it is not an empty route, one last edge is added to connect the final delivery back to home base
        if previous_index is not node_index:
            G.add_edge(node_index, 0, #length=routing.GetArcCostForVehicle(node_index, 0, vehicle_id),
                       color=colour_finder(vehicle_id))
            edges[(node_index, 0)] = routing.GetArcCostForVehicle(node_index, 0, vehicle_id)

        my_layout = nx.spring_layout(G)  # Set layout for edge labels and graph

        # this route graph is drawn
        if len(G) is not 0:
            current_plot += 1
            fig.add_subplot(1,plot_num,current_plot) # add subplot to a 1x3 matrix
            nx.draw(G,
                    node_color=colour_finder(vehicle_id),
                    pos=my_layout,
                    node_size=200,
                    with_labels=True,
                    font_size=10)
            # with edge labels added, which are the "cost" parameter for that segment of the route
            nx.draw_networkx_edge_labels(G, edge_labels=edges, pos=my_layout, font_size=8)
            #plt.show()
        G.clear()


def colour_finder(vehicle):
    """This function is used to get the graph color, where vehicle is the
    vehicle index for the given data set. Returns a color in string form."""

    switcher = {
        0: "blue",
        1: "red",
        2: "green",
        3: "yellow",
        4: "purple",
        5: "brown"
    }
    return switcher.get(vehicle, "Invalid month")


if __name__ == '__main__':

    ###################################### TESTING AREA ######################################

    # Instantiate the data problem using the MapQuest library
    #data = create_data_model(num_vehicles=1)
    
    # Instantiate the data problem using the Pickle files
    data = loadPickle(fileName='CostMatrix-1_02-1.pickle')
    #data = loadPickle(fileName='CostMatrix-1_03-1.pickle')
    #data = loadPickle(fileName='CostMatrix-1_07-1.pickle')
    #data = loadPickle(fileName='CostMatrix-2_18-1.pickle')

    print("\n----------------- TRAVELLING SALESMAN PROBLEM -----------------\n")
    # Perform TSP with 1 vehicle & no capacity limits
    data['num_vehicles'] = 1
    tsp_processing(data=data)

    print("\n----------------- VEHICLE ROUTING PROBLEM -----------------\n")
    # Perform VRP with 4 vehicles & no capacity limits
    data['num_vehicles'] = 5
    vrp_processing(data=data)

    print("\n----------------- LIMITED CAPACITY VEHICLE ROUTING PROBLEM ----------------- \n")
    # Perform VRP with 4 vehicles w/ limited capacity (Can change to 1 vehicle to do TSP w/ capacity)
    data['num_vehicles'] = 5
    
    data['demands'] = loadPickle(fileName='LoadMatrix-1_02-1.pickle') # Load weights for the deliveries
    #data['demands'] = loadPickle(fileName='LoadMatrix-1_03-1.pickle') # Load weights for the deliveries
    #data['demands'] = loadPickle(fileName='LoadMatrix-1_07-1.pickle') # Load weights for the deliveries
    #data['demands'] = loadPickle(fileName='LoadMatrix-2_18-1.pickle') # Load weights for the deliveries
    
    data['vehicle_capacities'] = [3000, 3000, 3000, 3000, 3000]
    vrp_limited_processing(data=data)

    """Show complete delivery destination graph"""
    #graph_and_edges = create_graph_from_mapquest()
    #routeGraph = graph_and_edges[0]
    #edges = graph_and_edges[1]
    #display_graph(routeGraph, edges)  

    ###################################### EXPERIMENTS LIST ######################################

    # ------------ BASELINE ALGORITHM ------------
    # 1.) Randomly Chosen [All Pickles]


    # ------------ PATH_CHEAPEST_ARC ALGORITHM w/ GREEDY DESCENT LOCAL SEARCH ------------
    # 1.) TSP(Unlimited Capacity) VS. VRP(Unlimited Capacity) [All Pickles]
    # 2.) TSP(Unlimited Capacity) VS. VRP(Limited Capacity) [All Pickles]
    # 3.) TSP(Limited Capacity) VS. VRP (Limited Capacity) [All Pickles]


    # ------------ SAVINGS ALGORITHM w/ GREEDY DESCENT LOCAL SEARCH ------------
    # 1.) TSP(Unlimited Capacity) VS. VRP(Unlimited Capacity) [All Pickles]
    # 2.) TSP(Unlimited Capacity( VS. VRP(Limited Capacity) [All Pickles]
    # 3.) TSP(Limited Capacity) VS. VRP (Limited Capacity) [All Pickles]

